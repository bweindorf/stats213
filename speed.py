import subprocess
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import json

t_start = time.time()
currenttime = t_start

wb = Workbook()
ws = wb.create_sheet("Baseline Test")
ws.cell(row = 1, column = 1, value = "Time Stamp")
ws.cell(row = 1, column = 2, value = "Relative Time (s)")
ws.cell(row = 1, column = 3, value = "Sample Number")
ws.cell(row = 1, column = 4, value = "Ping (ms)")
ws.cell(row = 1, column = 5, value = "Upload (Mbps)")
ws.cell(row = 1, column = 6, value = "Download (Mbps)")
ws.cell(row = 1, column = 7, value = "Temperature (F)")
ws.cell(row = 1, column = 8, value = "Windspeed (mph)")
i = 1

weather = subprocess.Popen('curl https://api.weather.gov/gridpoints/CTP/102,31/forecast', shell=True, stdout=subprocess.PIPE)
stdout, stderr = weather.communicate()
stdout = json.loads(stdout.decode("utf-8"))
windspeed = stdout["properties"]["periods"][0]["windSpeed"]
temp = stdout["properties"]["periods"][0]["temperature"]

windspeed = windspeed.split(" ")
if len(windspeed) == 4:
    windspeed = (1/2)*(float(windspeed[0])+float(windspeed[2]))
else:
    windspeed = float





ws.cell(row = 2, column = 7, value = temp)
ws.cell(row = 2, column = 8, value = windspeed)



while currenttime - t_start < 10800:
    currenttime = time.time()
    now = datetime.datetime.now()
    try:
        i+=1
        data = subprocess.check_output("speedtest-cli --simple", shell=True).decode("utf-8")
        data = data.splitlines()

        ping = data[0].split(" ")[1]
        upload = data[1].split(" ")[1]
        download = data[2].split(" ")[1]


        ws.cell(row = i, column = 1, value = now.strftime("%H:%M:%S"))
        ws.cell(row = i, column = 2, value = (currenttime - t_start))
        ws.cell(row = i, column = 3, value = i - 1)
        ws.cell(row = i, column = 4, value = float(ping))
        ws.cell(row = i, column = 5, value = float(upload))
        ws.cell(row = i, column = 6, value = float(download))

        wb.save("Baseline5xlsx")

    except:
        continue
